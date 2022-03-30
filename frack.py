#!/bin/python
# -*- coding: utf-8 -*-
#########################################################################################################
# Remember to export the location of the service account key!                                           #
# export GOOGLE_APPLICATION_CREDENTIALS=/home/stingray/TestingBigQuery-1209d661662d.json                #
#########################################################################################################
# TODO: DELETE FROM `testingbigquery-306308.Breach_Data.Breach_Data` WHERE site IS NULL AND breach IS NULL
# TODO: Add bulk parsing
#########################################################################################################
# Examples: ./frack.py parse -p -i wattpad_24133700_lines.txt -y 2021 -n None -w wattpad.com            #
# ./frack.py parse -p -y 2019 -n Collection#1 -w 3dsiso.com -d -u -i Collection#1_3DSISO.com_2019.csv   #
#########################################################################################################

import sys

if sys.version_info < (3, 9):
    print('Please upgrade your Python version to 3.9.0 or higher')
    sys.exit()

import argparse, csv, importlib
import os, sys, time, io, pyorc, re
from string import ascii_uppercase

from google.cloud import bigquery
from google.cloud import storage
from google.cloud.exceptions import NotFound
from hurry.filesize import size, verbose
from openpyxl import Workbook
from tabulate import tabulate

currentlocation = os.path.dirname(os.path.abspath(__file__))

# Change this to reflect your creds file or just rename your creds file.
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = currentlocation + "/creds.json"

# Change these to match your environment.
project_name = "testingbigquery-306308"
bucket_name = "ingestionbucket_frack"
bucket_uri = "gs://ingestionbucket_frack/*.orc"

# You can leave this one. The tool will create these for you on the first ingestion.
# table_id = 'Breach_Data.Part_Data'
table_id = 'Breach_Data.Breach_Data'

# static structure for pyorc files
pyorc_struct = ("struct<"
                "breach:string,site:string,year:int,domain:string,email:string,"
                "password:string,hash:string,salt:string"
                ">")

class txtcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def main():
    splash()
    parser = argparse.ArgumentParser(prog='frack.py', usage='%(prog)s ',
                                     description=('This is a tool to help you manage and query '
                                                  'breach data stored in a BigQuery table.'))
    subparsers = parser.add_subparsers()
    parser.add_mutually_exclusive_group(required=False)
    parser_query = subparsers.add_parser('query')
    parser_query.set_defaults(func=search)
    parser_query.add_argument('-i', '--inputfile', help="Input file containing domains.")
    parser_query.add_argument('-d', '--singledomain', help="Specify a single domain to search for.")
    parser_db = subparsers.add_parser('db')
    parser_db.set_defaults(func=maintain)
    parser_db.add_argument('-c', '--count', action='store_true', help="Count the lines in the dataset.")
    parser_db.add_argument('-n', '--nomnom', action='store_true',
                           help="Trigger the ingestion of the ORC files in the ingesting bucket.")
    parser_db.add_argument('-d', '--delete', action='store_true',
                           help="After ingestion clean the bucket by removing all the .orc files.")
    parser_db.add_argument('-t', '--top', help="Display the top <n> passwords in the database.")
    parser_db.add_argument('-w', '--web', action='store_true',
                           help="Display the websites with line count in the database.")
    parser_db.add_argument('-b', '--breach', action='store_true',
                           help="Display the breaches with line count in the database.")
    parser_db.add_argument('-f', '--file', action='store_true',
                           help="Save the output from these queries in an Excel sheet.")
    parser_parse = subparsers.add_parser('parse')
    parser_parse.set_defaults(func=parse)
    parser_parse.add_argument('-i', '--inputfile', help="File to import from. Default Format: e-mail:hash",
                              required=True)
    parser_parse.add_argument('-m', '--module', help="The parser module to use")
    parser_parse.add_argument('-y', '--year', help="Year the breach was released.")
    parser_parse.add_argument('-n', '--name', help="Name of the breach ex. Collection#1.")
    parser_parse.add_argument('-w', '--website', help="Website that was breached.")
    parser_parse.add_argument('-p', '--passwords', action='store_true',
                              help="Use if file contains passwords instead of hashes ex. e-mail:password.")
    parser_parse.add_argument('-s', '--salt', action='store_true',
                              help="Use if file contains hashes with salts ex. e-mail:hash:salt.")
    parser_parse.add_argument('-d', '--nodel', action='store_false', help="Don't delete the error file. Only available on CSV import not when using plugins!")
    parser_parse.add_argument('-u', '--upload', action='store_true',
                              help="Upload the file to the ingestion bucket after parsing.")

    if not len(sys.argv) > 1:
        parser.print_help()
        sys.exit()
    args = parser.parse_args()
    args.func(args)


#########################################################################################################
# Query the breaches and websites in the database                                                       #
#########################################################################################################
def stats(what, export):
    sql_query = ""
    if what == "web":
        sql_query = """
        SELECT DISTINCT site, Count(site) as number,
        FROM `""" + project_name + '.' + table_id + """` 
        GROUP BY site ORDER BY number desc
        """
    elif what == "breach":
        sql_query = """
        SELECT DISTINCT breach, Count(breach) as number,
        FROM `""" + project_name + '.' + table_id + """` 
        GROUP BY breach ORDER BY number desc 
        """

    print("Querying Breach Database...")
    print("Note: This may take a while")
    client = bigquery.Client()
    tic = time.perf_counter()
    query_job = client.query(sql_query)
    # Wait for the job to complete
    results = query_job.result()
    toc = time.perf_counter()
    print(f"Query completed in {toc - tic:0.4f} seconds")

    # Convert the results to a pandas dataframe and save it to a .csv file.
    df = query_job.to_dataframe()
    print(tabulate(df, headers='keys', tablefmt='psql'))

    if not export:
        return

    print("Flushing data to file...")
    df.to_csv("temp.csv", index=False, header=False)
    wb = Workbook()
    ws1 = wb.active
    if what == "web":
        ws1.title = "Websites in the DB"
        ws1['A1'] = "Website"
        ws1['B1'] = "Lines"
        dest_filename = "Websites_in_DB.xlsx"
    elif what == "breach":
        ws1.title = "Breaches in the DB"
        ws1['A1'] = "Breach"
        ws1['B1'] = "Lines"
        dest_filename = "Breaches_in_DB.xlsx"

    for column in ascii_uppercase:
        if column == 'A':
            ws1.column_dimensions[column].width = 30
        elif column == 'B':
            ws1.column_dimensions[column].width = 30
        else:
            ws1.column_dimensions[column].width = 15

    try:
        with open('temp.csv', 'r') as f:
            for row in csv.reader(f):
                ws1.append(row)
        wb.save(dest_filename)
        print('File written to: ./' + dest_filename)
        os.remove("temp.csv")
    except:
        print("Something bad happened while creating the Excel sheet!")


#########################################################################################################
# Query the top <n> passwords in the current database                                                   #
#########################################################################################################
def show_top_passwords(size, export):
    print("Looking up the top " + size + " passwords...")

    sql_top = """
    SELECT DISTINCT password, COUNT(password) as used
    FROM `""" + project_name + '.' + table_id + """` 
    GROUP BY password ORDER BY used desc limit """ + size + """
    """

    print("Querying Breach Database...")
    print("Note: This may take a while")
    client = bigquery.Client()
    tic = time.perf_counter()
    query_job = client.query(sql_top)
    # Wait for the job to complete
    results = query_job.result()
    toc = time.perf_counter()
    print(f"Query completed in {toc - tic:0.4f} seconds")

    # Convert the results to a pandas dataframe and save it to a .csv file.
    df = query_job.to_dataframe()
    print(tabulate(df, headers='keys', tablefmt='psql'))

    if not export:
        return

    print("Flushing data to file...")
    df.to_csv("temp.csv", index=False, header=False)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Top " + size + " Passwords"
    ws1['A1'] = "Password"
    ws1['B1'] = "Count"

    for column in ascii_uppercase:
        if (column == 'A'):
            ws1.column_dimensions[column].width = 20
        elif (column == 'B'):
            ws1.column_dimensions[column].width = 20
        else:
            ws1.column_dimensions[column].width = 15

    dest_filename = "Top_" + size + "_Passwords.xlsx"
    try:
        with open('temp.csv', 'r') as f:
            for row in csv.reader(f):
                ws1.append(row)
        wb.save(dest_filename)
        print('File written to: ./' + dest_filename)
        os.remove("temp.csv")
    except:
        print("Something bad happened while creating the Excel sheet!")


#########################################################################################################
# Convert bytes to human readable form                                                                  #
#########################################################################################################
def convert_bytes(num):
    for x in ['bytes', 'KB', 'MB', 'GB', 'TB']:
        if num < 1024.0:
            return "%3.1f %s" % (num, x)
        num /= 1024.0


#########################################################################################################
# Easy way to display the file size in human readable form                                              #
#########################################################################################################
def file_size(file_path):
    if os.path.isfile(file_path):
        file_info = os.stat(file_path)
        return convert_bytes(file_info.st_size)


#########################################################################################################
# Uploads the new .orc file to the ingestion bucket                                                     #
#########################################################################################################
def upload_blob(bucket_name, source_file_name):
    print(txtcolors.OKGREEN + "Uploading File to storage bucket..." + txtcolors.ENDC)
    storage_client = storage.Client()
    bucket = storage_client.get_bucket(bucket_name)
    blob = bucket.blob(source_file_name)
    blob.upload_from_filename(source_file_name)
    print(txtcolors.OKGREEN + f"File uploaded to the {bucket_name} storage bucket." + txtcolors.ENDC)


#########################################################################################################
# Validate the data in the row                                                                          #
#########################################################################################################
def validate_data(row, passbool):
    # Regex for validating an E-Mail address
    #email_regex = '^(\w|\.|\_|\-)+[@](\w|\_|\-|\.)+[.]\w{2,3}$'
    email_regex = '^[a-zA-Z0-9.!#$%&’*+/=?^_`{|}~-]+@[a-zA-Z0-9-]+(?:\.[a-zA-Z0-9-]+)*$'

    hash = domain = email = salt = password = ''

    if passbool:
        email = row[0].strip()
        password = row[1].strip()
        hash == ''
    else:
        email = row[0].strip()
        hash = row[1].strip()
        try:
            salt = row[2].strip()
        except:
            salt == ''

    try:
        domain = row[0].split('@')[1]
    except:
        return False

    if password == '' and hash == '':
        return False

    if domain == '':
        return False

    if ( not re.search(email_regex, email)):
        return False

    return True

#########################################################################################################
# Parse a file to .orc format for upload to the ingestion bucket. If a parse module is defined, use     #
# that module to parse the file to .orc                                                                 #
#########################################################################################################
def parse(args):
    if args.module:
        try:
            mod_src = importlib.import_module(f'parsers.{args.module}')
        except ImportError as e:
            print(txtcolors.FAIL + f'unable to import parser with error: {e}' + txtcolors.ENDC)
            return

        parser = mod_src.Parse(args.inputfile, args.upload)
        parser.process()

    elif (args.year is None) or (args.name is None) or (args.website is None):
        print(txtcolors.FAIL + "The arguments for name, year and website is required!" + txtcolors.ENDC)
    else:
        print(txtcolors.OKGREEN + "Parsing the .csv file..." + txtcolors.ENDC)
        errorcount = 0
        writecount = 0
        mytup = []
        domain = ''
        destination = f'{args.year}.{args.website}.{args.name}.orc'
        errors_file = f'{args.year}.{args.website}.errors'

        with open(args.inputfile, "r") as csvfile:
            reader = csv.reader((line.replace('\0', '') for line in csvfile), delimiter=',')
            error_file = open(errors_file, 'w')
            error_writer = csv.writer(error_file)
            with open(destination, 'wb') as data:
                with pyorc.Writer(data, pyorc_struct) as writer:
                    for row in reader:
                        if validate_data(row, args.passwords):
                            domain = row[0].split('@')[1]
                            if args.passwords:
                                mytup = [args.name, args.website, int(args.year), domain, row[0], row[1], '', '']
                            elif args.salt:
                                mytup = [args.name, args.website, int(args.year), domain, row[0], '', row[1], row[2]]
                            else:
                                mytup = [args.name, args.website, int(args.year), domain, row[0], '', row[1], '']
                            writer.write(tuple(mytup))
                            writecount += 1
                        else:
                            errorcount += 1
                            error_writer.writerow(row)
                writecount_formatted = '{:,}'.format(writecount)
                errorcount_formatted = '{:,}'.format(errorcount)
                sys.stdout.write("\rGood Lines: %s Bad Lines: %s" % (writecount_formatted, errorcount_formatted))
                sys.stdout.flush()
            sys.stdout.write('\n')
            error_file.close()
            print(txtcolors.OKBLUE + "Size of import file: " + file_size(args.inputfile) + txtcolors.ENDC)
            print(txtcolors.OKGREEN + f'File written to: {destination}'+ txtcolors.ENDC)
            print(txtcolors.OKBLUE + "ORC Size: " + file_size(destination) + txtcolors.ENDC)
            if args.nodel:
                os.remove(errors_file)
            if args.upload:
                print(txtcolors.OKGREEN + "Uploading .orc to bucket ..." + txtcolors.ENDC)
                upload_blob(bucket_name, destination)

#########################################################################################################
# Perform the DB maintenance commands specified in args.                                                #
#########################################################################################################
def maintain(args):
    # If the count arg is passed, do just that and exit.
    if args.count:
        count_dataset()
    # If the nomnom arg is passed, start the ingestion job.
    if args.nomnom:
        ingest_orc(args.delete)
    if args.top:
        show_top_passwords(args.top, args.file)
    if args.web:
        stats("web", args.file)
    if args.breach:
        stats("breach", args.file)


#########################################################################################################
# Perform the search on the Database.                                                                   #
#########################################################################################################
def search(args):
    domains = []
    # If there's an input file, read the contents and populate domains.
    if args.inputfile:
        domains = [l.strip() for l in open(args.inputfile).readlines()]
    # If there's a single domain, strip it and set it as domains.
    if args.singledomain:
        domains = [args.singledomain.strip()]

    sql_search = """
    SELECT *
    FROM `""" + project_name + '.' + table_id + """` 
    WHERE UPPER(domain) = UPPER(""" + '"' + '") OR UPPER(domain)=UPPER("'.join(domains) + '")' + """
    """
    print("Querying Breach Database...")
    client = bigquery.Client()
    tic = time.perf_counter()
    query_job = client.query(sql_search)
    # Wait for the job to complete
    results = query_job.result()
    toc = time.perf_counter()
    print(f"Query completed in {toc - tic:0.4f} seconds")
    print("{:,}".format(results.total_rows) + " lines in the dataset")

    # Convert the results to a pandas dataframe and save it to a .csv file.
    df = query_job.to_dataframe()
    print("Flushing data to file...")
    df.to_csv("temp.csv", index=False, header=False)
    df.query("password != ''", inplace=True)
    passwords = df['password'].tolist()
    create_excel("temp.csv", args.singledomain, passwords)

    # Clean up
    try:
        os.remove('temp.csv')
    except:
        print("Error cleaning up!")


#########################################################################################################
# Display the splash screen.                                                                            #
#########################################################################################################
def splash():
    print('')
    print(txtcolors.OKGREEN + ' ███████╗██████╗░░█████╗░░█████╗░██╗░░██╗' + txtcolors.ENDC)
    print(txtcolors.OKGREEN + ' ██╔════╝██╔══██╗██╔══██╗██╔══██╗██║░██╔╝' + txtcolors.ENDC)
    print(txtcolors.OKGREEN + ' █████╗░░██████╔╝███████║██║░░╚═╝█████═╝░' + txtcolors.ENDC)
    print(txtcolors.OKGREEN + ' ██╔══╝░░██╔══██╗██╔══██║██║░░██╗██╔═██╗░' + txtcolors.ENDC)
    print(txtcolors.OKGREEN + ' ██║░░░░░██║░░██║██║░░██║╚█████╔╝██║░╚██╗' + txtcolors.ENDC)
    print(txtcolors.OKGREEN + ' ╚═╝░░░░░╚═╝░░╚═╝╚═╝░░╚═╝░╚════╝░╚═╝░░╚═╝' + txtcolors.ENDC)
    print(txtcolors.OKGREEN + '    - By: William Vermaak ' + txtcolors.ENDC)
    print('')


#########################################################################################################
# Count the lines in the current dataset.                                                               #
#########################################################################################################
def count_dataset():
    client = bigquery.Client()
    try:
        destination_table = client.get_table(table_id)
        print("Current DB Consists of {:,} rows.".format(destination_table.num_rows))
        print("Current DB Size {}".format(size(destination_table.num_bytes, system=verbose)))
    except NotFound:
        print("DB has not been created yet. Please run frack.py db -n to create a blank DB.")
    sys.exit()


#########################################################################################################
# Ingest all of the .orc files in the ingestion_bucket.                                                 #
#########################################################################################################
def ingest_orc(delete):
    client = bigquery.Client()
    schema = [
        bigquery.SchemaField("breach", "STRING"),
        bigquery.SchemaField("site", "STRING"),
        bigquery.SchemaField("year", "INTEGER"),
        bigquery.SchemaField("domain", "STRING"),
        bigquery.SchemaField("email", "STRING"),
        bigquery.SchemaField("password", "STRING"),
        bigquery.SchemaField("hash", "STRING"),
        bigquery.SchemaField("salt", "STRING"),
    ]
    try:
        destination_table = client.get_table(table_id)
    except NotFound:
        print("Table not found. Creating a blank dataset...")
        dataset_id = "{}.Breach_Data".format(client.project)
        dataset = bigquery.Dataset(dataset_id)
        dataset.location = "US"
        dataset = client.create_dataset(dataset, timeout=30)  # API Request
        print("Creating blank table...")
        table = bigquery.Table(project_name + '.' + table_id, schema=schema)
        table = client.create_table(table)  # API Request
        destination_table = client.get_table(table_id)
    print("Current DB Consists of {:,} rows.".format(destination_table.num_rows))
    job_config = bigquery.LoadJobConfig(
        schema=schema,
    )
    body = io.BytesIO(b",,,,,,,")
    #body = six.BytesIO(b",,,,,,,")
    #body = b",,,,,,,"
    client.load_table_from_file(body, table_id, job_config=job_config).result()
    previous_rows = client.get_table(table_id).num_rows
    assert previous_rows > 0
    job_config = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        source_format=bigquery.SourceFormat.ORC
    )

    load_job = client.load_table_from_uri(  # API Request
        bucket_uri, table_id, job_config=job_config
    )
    tic = time.perf_counter()
    print("Job has been created :) waiting for it to finish.")
    print("This may take a while depending on the amount of data to ingest.")
    # Waits for the job to complete.
    load_job.result()
    toc = time.perf_counter()
    print(f"Ingestion completed in {toc - tic:0.4f} seconds")
    print("Job has finished.")
    destination_table = client.get_table(table_id)
    print("After Update DB Consists of {:,} rows.".format(destination_table.num_rows))
    if delete:
        empty_bucket(bucket_name)
    sys.exit()


#########################################################################################################
#
#########################################################################################################
def empty_bucket(local_bucket_name):
    storage_client = storage.Client()
    bucket = storage_client.bucket(local_bucket_name)
    print("Listing Files")
    blobs = bucket.list_blobs()
    for blob in blobs:
        if blob.name.endswith('.orc'):
            print(blob.name + ' -> Deleted')
            blob.delete()


#########################################################################################################
# Create the Excel sheet with the results to make it easy to filter and to send to customers.           #
#########################################################################################################
def create_excel(csv_name, singledomain, passwords):
    unique_pass = []
    topTen = []
    clean_pass = []

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Breach Data"
    ws1['A1'] = "Breach"
    ws1['B1'] = "Website"
    ws1['C1'] = "Year"
    ws1['D1'] = "Domain"
    ws1['E1'] = "E-Mail"
    ws1['F1'] = "Password"
    ws1['G1'] = "Hash"
    ws1['H1'] = "Salt"
    ws2 = wb.create_sheet(title="Top 10 Passwords")
    ws2['A1'] = "Password"
    ws2['B1'] = "Count"
    ws3 = wb.create_sheet(title="Unique Passwords")
    ws3['A1'] = "Password"

    for column in ascii_uppercase:
        if column == 'A':
            ws1.column_dimensions[column].width = 20
            ws2.column_dimensions[column].width = 30
            ws3.column_dimensions[column].width = 30
        elif column == 'B':
            ws1.column_dimensions[column].width = 20
            ws2.column_dimensions[column].width = 15
        elif column == 'C':
            ws1.column_dimensions[column].width = 10
        elif column == 'D':
            ws1.column_dimensions[column].width = 35
        elif column == 'E':
            ws1.column_dimensions[column].width = 25
        elif column == 'F':
            ws1.column_dimensions[column].width = 30
        elif column == 'G':
            ws1.column_dimensions[column].width = 10
        else:
            ws1.column_dimensions[column].width = 15

    print("Performing the password analysis...")

    # Remove None values from tuple.
    for val in passwords:
        if val is not None:
            clean_pass.append(val)

    # Add a count to the passwords in a tuple so we can use it later.
    topTen = [item for item in Counter(clean_pass).most_common()]

    # Use the first 10 items in our tuple to write in the sheet.
    for i in range(min(10, len(topTen))):
        cellA = 'A' + str(i + 2)
        cellB = 'B' + str(i + 2)
        count, word = topTen[i]
        ws2[cellA] = word
        ws2[cellB] = count

    # Get a list of unique passwords.
    for i in clean_pass:
        if i not in unique_pass:
            unique_pass.append(i)

    # Populate worksheet 3 with all the unique passwords for the domain.
    unique_pass.sort()
    for i in range(len(unique_pass)):
        cellA = 'A' + str(i + 2)
        ws3[cellA] = unique_pass[i]

    # If we're working with a single domain, rename the Excel sheet to that domain for
    # easier reference. If it's a list we'll just call it Breach_Data.xlsx
    if singledomain:
        dest_filename = 'Breach_Data_' + singledomain.strip() + '.xlsx'
    else:
        dest_filename = 'Breach_Data.xlsx'

    try:
        with open('temp.csv', 'r') as f:
            for row in csv.reader(f):
                ws1.append(row)
        wb.save(dest_filename)
        print('File written to: ./' + dest_filename)
    except:
        print("Something bad happened while creating the Excel sheet!")


if __name__ == "__main__":
    main()
